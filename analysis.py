import openpyxl
import json

# Load Excel
wb = openpyxl.load_workbook('json/BOD Scores History_2024-08-31.xlsx')

# Champions sheet
champions_ws = wb['Champions']
excel_tournaments = []
for row in champions_ws.iter_rows(min_row=2, values_only=True):
    if row[0]:  # Has date
        excel_tournaments.append({
            'date': str(row[0])[:10] if row[0] else None,
            'format': row[1],
            'bod_num': row[2],
            'champion': row[3]
        })

print(f"=== Excel Champions Sheet ===")
print(f"Total tournaments: {len(excel_tournaments)}")
print(f"BOD numbers range: {min(t['bod_num'] for t in excel_tournaments if t['bod_num'])} to {max(t['bod_num'] for t in excel_tournaments if t['bod_num'])}")

# All Scores sheet
scores_ws = wb['All Scores']
excel_scores = []
excel_score_dates = set()
for row in scores_ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        date_str = str(row[0])[:10] if row[0] else None
        excel_score_dates.add(date_str)
        excel_scores.append({'date': date_str, 'format': row[1]})

print(f"\n=== Excel All Scores Sheet ===")
print(f"Total score entries: {len(excel_scores)}")
print(f"Unique dates: {len(excel_score_dates)}")

# JSON Scores
with open('json/All Scores.json', 'r') as f:
    json_scores = json.load(f)

json_dates = set(entry.get('Date', '') for entry in json_scores)
print(f"\n=== JSON All Scores ===")
print(f"Total entries: {len(json_scores)}")
print(f"Unique dates: {len(json_dates)}")

# Compare
missing_in_json = excel_score_dates - json_dates
extra_in_json = json_dates - excel_score_dates

print(f"\n=== Date Comparison ===")
print(f"Dates in Excel only: {sorted(missing_in_json) if missing_in_json else 'None'}")
print(f"Dates in JSON only: {sorted(extra_in_json) if extra_in_json else 'None'}")

# Check tournament JSON files
import os
json_files = [f for f in os.listdir('json') if f.endswith('.json') and f not in ['All Scores.json', 'All Players.json', 'Champions.json']]
print(f"\n=== Individual Tournament JSON Files ===")
print(f"Total individual tournament files: {len(json_files)}")

# Parse dates from filenames
file_dates = set()
for f in json_files:
    parts = f.replace('.json', '').split(' ')
    if len(parts) >= 1 and '-' in parts[0]:
        file_dates.add(parts[0])

print(f"Unique dates from filenames: {len(file_dates)}")

# Compare with Excel champions
excel_champ_dates = set(t['date'] for t in excel_tournaments if t['date'])
missing_files = excel_champ_dates - file_dates
print(f"\nTournaments in Excel Champions without JSON file: {len(missing_files)}")
if missing_files:
    print(f"  Missing: {sorted(missing_files)[:10]}")
